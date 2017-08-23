#! /usr/bin/env python
# _*_ coding:utf-8 _*_

from urllib import request
import http.cookiejar
import json,re
import csv
#读取excel使用(支持03)  
import xlrd  
#写入excel使用(支持03)  
import xlwt 
#读取execel使用(支持07)  
from openpyxl import Workbook  
#写入excel使用(支持07)  
from openpyxl import load_workbook  

#定义获取内容的函数
def getContent(start,limit):
    #内容URL
    contentURL='http://218.204.155.167:8081/cms/farmers.do?method=query&_dc=1&start='+str(start)+'&limit='+str(limit)+'&twnId=&vlgId=&grpId=&farCardsn=&farSn=&farName=&valid=1&hclId=&hatId=&famIsAssistance=&jatId=&frlId=&natId=&sexId=&h_rtiId=&istId=&farNewBorn=&startDate=2017&htrId=&birtStartDate=&birtEndDate=&registerdateBegin=&registerdateEnd=&lastValid=&serialId=&fjoPaymentmode=&farCardid=&payStartDate=&payEndDate=&qFlag=0&sid=NUEplkcfFJcWM2K8fYIIbw%3D%3D'

    req=request.Request(contentURL,postData,requestHeaders)
    content=request.urlopen(req).read().decode('utf-8')
    #清洗json数据
    jsonData=re.findall(r'root:(.*)}', content)
    jsonData='{"root":'+jsonData[0]+"}"
    data=json.loads(jsonData)
    content=data['root']
    return content

#登陆POST数据发送的头部信息
requestHeaders={
    'Host': '218.204.155.167:8081',
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.10; rv:41.0) Gecko/20100101 Firefox/41.0',
    'Connection': 'keep-alive',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Referer': 'http://218.204.155.167:8081/cms/join/joinList.jsp',
    'Cookie': 'ys-userName=s%3A36072104001; JSESSIONID=abcPJdNu2kh4xhncE0rQv',
    'Pragma': 'no-cache',
    'Cache-Control': 'no-cache',
    'Content-Length': '360'
}
postData={
    
}
#表格中文
table_title=['个人编码','医疗证号','地址','姓名','家庭关系','性别','身份证','出生日期','参合状态','救助类型','参合时间','缴费时间','缴费方式','缴费经办人','住院补偿金额','门诊大病补偿金额','门诊补偿金额','特殊补偿金额','其他补偿金额(定额)','家庭账户余额','大病保险补偿', '参合属性','户属性','户口类别','救助家庭','民族','联系电话','联系手机','移动手机','电信手机','联通手机','银行账号','退费状态','家庭编码','所属乡镇','所属村','所属组','参合属性','户属性','户口类别','救助家庭','民族','性别','救助人员','救助类型','家庭人口数','家庭农业人口数','家庭现住农业人口数','当前状态','家庭关系','新生儿','审批时间','锁定时间','是否发卡','退费情况','迁出状态']
#表格对应字段
table_field=['serialId','farCardsn','fullName','farName','frlName','sexName','farSn','farBirthday','valid','rtiName','fjoJoindate','fjoPaymentdate','fjoPaymentmode','fjoPaymenterName','fjoHoscommoney','fjoExtracommoney','fjoCliniccommoney','fjoNotophoscommoney','fjoOtherhoscommoney','facNowaccountmoney','fjoInsurancecommoney','jatName','hatName','hclName','famIsAssistanceName','natName','farTel','farMtel','farMtelm','farMtelt','farMtelu','farBanksn','fjoIsretfee','famId','twnId', 'vlgId','areaSid','jatId','hatId','hclId','famIsassistance','natId','sexId','farIsassistance', 'rtiId','famPerson','famAgriculturalperson','famLiveperson','istId','frlId','fjoNewborn','farJoindate', 'fjoRegisterdate','farCardid','isretfee','htrId']
# excel 参数设置
current_row=2
start_colum=1
page_size=500
# 写入文件名称
writepath=r"/Users/leo/trscases/gitproject/StockSpider/medicalInfomation_jb01.xlsx" 
#写表头
wb=Workbook()  
sheet=wb.create_sheet("医保数据",0) 

for value in table_title:
    sheet.cell(row=1,column=start_colum).value=value
    start_colum+=1

currentRow=2

# 从第1页到396页
for i in range(0,1):
    # 重新打开excel
    if(i%5==0):
        wb.save(writepath)
        wb=load_workbook(writepath)   
        sheet=wb.get_sheet_by_name("医保数据")  
    content=getContent(32500,page_size)
    for o in content:
        startColum=1
        for i in table_field:
            try:
                sheet.cell(row=currentRow,column=startColum).value=o[i]
            except Exception as e:
                sheet.cell(row=currentRow,column=startColum).value=''
            startColum+=1
        currentRow+=1
wb.save(writepath)
